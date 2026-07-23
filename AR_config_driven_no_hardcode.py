# ==========================================================
# IMPORTS
# ==========================================================

import os
import sys
import time
import shutil
import logging
import traceback
import platform
import pyotp
import pandas as pd
import openpyxl
import json
import threading
import atexit
import re
import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# Optional desktop notification libraries - degrade gracefully if missing
try:
    from plyer import notification as _plyer_notification
    _DESKTOP_NOTIFY_BACKEND = "plyer"
except ImportError:
    try:
        from win10toast import ToastNotifier as _Win10Toast
        _DESKTOP_NOTIFY_BACKEND = "win10toast"
    except ImportError:
        _DESKTOP_NOTIFY_BACKEND = None

from pathlib import Path
from datetime import datetime, timedelta
from dataclasses import dataclass, asdict
from typing import Optional, List, Callable, Any, Tuple, Dict
from enum import Enum
from functools import wraps

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
    ElementNotInteractableException,
    ElementClickInterceptedException
)

try:
    from webdriver_manager.chrome import ChromeDriverManager as WDM_ChromeDriverManager
    _WDM_AVAILABLE = True
except ImportError:
    _WDM_AVAILABLE = False

# ==========================================================
# CONFIGURATION
# ==========================================================

class Config:
    """Centralized configuration management."""

    # URLs
    URL = "https://pvpm.practicevelocity.com/"

    # Browser
    HEADLESS = True
    WINDOW_SIZE = "1920,1080"
    DISABLE_IMAGES = True
    PAGE_LOAD_STRATEGY = "normal"   # normal/eager
    HEADLESS_STARTUP_DELAY = 1

    # Paths
    BASE_DIR = Path(r"C:\Users\mages\Desktop\Automation\AR Untouched")
    EXCEL_FILE = BASE_DIR / "Untouch.xlsx"
    DOWNLOAD_DIR = BASE_DIR / "Reports"
    LOG_DIR = BASE_DIR / "Logs"
    SCREENSHOT_DIR = LOG_DIR / "Screenshots"
    STATE_FILE = LOG_DIR / "automation_state.json"

    # Timeouts
    PAGE_LOAD_TIMEOUT = 180
    IMPLICIT_WAIT = 5
    DOWNLOAD_TIMEOUT = 180
    POPUP_TIMEOUT = 120
    ELEMENT_TIMEOUT = 60
    RETRY_DELAY = 3
    MAX_RETRIES = 3
    PRACTICE_SWITCH_TIMEOUT = 30

    # Report Types
    AGE24_INVOICE_SUFFIX = "AGE_24_ByInvcDate"
    AGE24_SERVICE_SUFFIX = "AGE24_Service"  # unused for filename now - Service Date report keeps its original downloaded filename
    FIN18_SUFFIX = "FIN18"

    # ------------------------------------------------------
    # Search results intermediate page
    # ------------------------------------------------------
    # Searching a report code (e.g. "AGE24") lands on a results LIST first -
    # the actual report (practice dropdown + date-type radio buttons) only
    # loads after clicking the matching link in that list. Map report code
    # -> the exact link text shown under "All Results". If a code isn't in
    # this map, the code falls back to clicking the FIRST result link,
    # which is normally correct since the search term is specific.
    REPORT_RESULT_LINK_TEXT = {
        "AGE24": "Non-Month End Aging",
        # "FIN18": "<exact link text from FIN18 search results>",  # fill in if the fallback ever picks the wrong one
    }

    # ------------------------------------------------------
    # Notifications
    # ------------------------------------------------------
    # Desktop toast notifications (uses plyer or win10toast if installed;
    # silently disabled if neither is available)
    DESKTOP_NOTIFICATIONS_ENABLED = True

    # Windows message box popup on run completion / fatal error (uses the
    # built-in ctypes Windows API, so no extra library is required). Shown
    # in a background thread so it never blocks cleanup/shutdown.
    MESSAGEBOX_ENABLED = True

    # Email notifications (set EMAIL_ENABLED = True and fill in the rest
    # to receive run summaries / failure alerts by email)
    EMAIL_ENABLED = False
    EMAIL_SMTP_HOST = "smtp.gmail.com"
    EMAIL_SMTP_PORT = 587
    EMAIL_USERNAME = ""          # sender address
    EMAIL_PASSWORD = ""          # app password / SMTP password
    EMAIL_FROM = ""              # defaults to EMAIL_USERNAME if blank
    EMAIL_TO = []                # list of recipient addresses, e.g. ["ar@yourclinic.com"]
    EMAIL_NOTIFY_ON_START = False
    EMAIL_NOTIFY_ON_SUCCESS = True
    EMAIL_NOTIFY_ON_FAILURE = True

    # ------------------------------------------------------
    # Circuit breaker
    # ------------------------------------------------------
    # If this many practices IN A ROW fail completely (all retries
    # exhausted), stop the run early instead of grinding through the
    # rest of the list against a dead session / down site.
    CIRCUIT_BREAKER_ENABLED = True
    CIRCUIT_BREAKER_THRESHOLD = 4

    # ------------------------------------------------------
    # Force reprocess
    # ------------------------------------------------------
    # When True, ignores the Excel "Status" column AND the saved
    # automation_state.json completed-list, and reprocesses every practice
    # in the sheet regardless of prior "Downloaded" status. Useful for
    # testing/re-runs against a sheet where Status is already filled in.
    # Leave False for normal day-to-day runs (so completed rows are skipped).
    FORCE_REPROCESS = False

    @classmethod
    def ensure_directories(cls):
        """Create all required directories."""
        for directory in [cls.DOWNLOAD_DIR, cls.LOG_DIR, cls.SCREENSHOT_DIR]:
            directory.mkdir(parents=True, exist_ok=True)

# ==========================================================
# ENUMS & DATA CLASSES
# ==========================================================

class Status(Enum):
    """Automation status states."""
    PENDING = "Pending"
    DOWNLOADED = "Downloaded"
    FAILED = "Failed"
    RETRY = "Retry"
    SKIPPED = "Skipped"
    PARTIAL = "Partial"  # New: Some reports downloaded, some failed

class ReportType(Enum):
    """Report type enumeration."""
    AGE24_INVOICE = "AGE24_Invoice"
    AGE24_SERVICE = "AGE24_Service"
    FIN18 = "FIN18"


class CircuitBreakerTripped(Exception):
    """Raised when too many consecutive practices fail, to stop a doomed run early."""
    pass

@dataclass
class Credentials:
    """Login credentials data class."""
    username: str
    password: str
    security_key: str

@dataclass
class AGE24Practice:
    """AGE24 practice data."""
    practice_type: str
    practice_name: str
    service_date: str
    status: str = ""
    row_number: int = 0
    invoice_downloaded: bool = False
    service_downloaded: bool = False

@dataclass
class FIN18Practice:
    """FIN18 practice data."""
    practice_type: str
    practice_name: str
    from_date: str
    to_date: str
    status: str = ""
    row_number: int = 0

@dataclass
class AutomationState:
    """Persistable automation state."""
    last_run: Optional[str] = None
    completed_practices: List[str] = None
    failed_practices: List[str] = None
    current_module: Optional[str] = None
    partial_practices: List[str] = None  # New

    def __post_init__(self):
        if self.completed_practices is None:
            self.completed_practices = []
        if self.failed_practices is None:
            self.failed_practices = []
        if self.partial_practices is None:
            self.partial_practices = []

# ==========================================================
# LOGGING SYSTEM
# ==========================================================

class ColoredFormatter(logging.Formatter):
    """Colored console output for logs."""

    COLORS = {
        'DEBUG': '\033[36m',      # Cyan
        'INFO': '\033[32m',       # Green
        'WARNING': '\033[33m',    # Yellow
        'ERROR': '\033[31m',      # Red
        'CRITICAL': '\033[35m',   # Magenta
        'RESET': '\033[0m'
    }

    def format(self, record):
        color = self.COLORS.get(record.levelname, self.COLORS['RESET'])
        reset = self.COLORS['RESET']
        record.levelname = f"{color}{record.levelname}{reset}"
        return super().format(record)


def setup_logging():
    """Configure advanced logging with file and console handlers."""

    Config.ensure_directories()

    log_file = Config.LOG_DIR / f"Automation_{datetime.now():%Y%m%d_%H%M%S}.log"

    # Root logger
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)

    # Clear existing handlers
    logger.handlers = []

    # File handler - detailed
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_format = logging.Formatter(
        "%(asctime)s | %(levelname)-8s | %(funcName)-20s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S"
    )
    file_handler.setFormatter(file_format)

    # Console handler - colored, less verbose
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    console_format = ColoredFormatter(
        "%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S"
    )
    console_handler.setFormatter(console_format)

    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


logger = setup_logging()


def log(message: str, level: str = "info"):
    """Unified logging function."""
    getattr(logger, level.lower(), logger.info)(message)

# ==========================================================
# STATE MANAGEMENT
# ==========================================================

class StateManager:
    """Manages persistent automation state for resume capability."""

    def __init__(self):
        self.state = AutomationState()
        self._load()

    def _load(self):
        """Load state from disk."""
        if Config.STATE_FILE.exists():
            try:
                with open(Config.STATE_FILE, 'r') as f:
                    data = json.load(f)
                    self.state = AutomationState(**data)
                log(f"State loaded: {len(self.state.completed_practices)} completed, {len(self.state.failed_practices)} failed")
            except Exception as e:
                log(f"Failed to load state: {e}", "warning")

    def save(self):
        """Save state to disk."""
        try:
            self.state.last_run = datetime.now().isoformat()
            with open(Config.STATE_FILE, 'w') as f:
                json.dump(asdict(self.state), f, indent=2)
        except Exception as e:
            log(f"Failed to save state: {e}", "warning")

    def mark_completed(self, practice_name: str):
        """Mark practice as completed."""
        if practice_name not in self.state.completed_practices:
            self.state.completed_practices.append(practice_name)
            self.save()

    def mark_failed(self, practice_name: str):
        """Mark practice as failed."""
        if practice_name not in self.state.failed_practices:
            self.state.failed_practices.append(practice_name)
            self.save()

    def mark_partial(self, practice_name: str):
        """Mark practice as partially completed."""
        if practice_name not in self.state.partial_practices:
            self.state.partial_practices.append(practice_name)
            self.save()

    def is_completed(self, practice_name: str) -> bool:
        """Check if practice was already completed."""
        return practice_name in self.state.completed_practices

    def is_failed(self, practice_name: str) -> bool:
        """Check if practice failed before."""
        return practice_name in self.state.failed_practices

    def set_module(self, module_name: str):
        """Set current module."""
        self.state.current_module = module_name
        self.save()

    def reset_practice(self, practice_name: str):
        """Reset practice status for reprocessing."""
        for lst in [self.state.completed_practices, self.state.failed_practices, self.state.partial_practices]:
            if practice_name in lst:
                lst.remove(practice_name)
        self.save()


state_manager = StateManager()

# ==========================================================
# NOTIFICATIONS
# ==========================================================

class NotificationManager:
    """Sends desktop toast and/or email notifications about run progress.

    Every method fails soft: a broken SMTP config or missing toast library
    will never crash the automation - it just logs a warning and moves on.
    """

    def __init__(self):
        self.desktop_available = (
            Config.DESKTOP_NOTIFICATIONS_ENABLED and _DESKTOP_NOTIFY_BACKEND is not None
        )
        if Config.DESKTOP_NOTIFICATIONS_ENABLED and not self.desktop_available:
            log("Desktop notifications requested but no backend found "
                "(pip install plyer OR win10toast to enable)", "warning")

    # ---------------- Desktop ----------------

    def send_desktop(self, title: str, message: str):
        if not self.desktop_available:
            return
        try:
            message = message[:250]  # toast backends truncate long text anyway
            if _DESKTOP_NOTIFY_BACKEND == "plyer":
                _plyer_notification.notify(
                    title=title, message=message, app_name="AR Automation", timeout=10
                )
            elif _DESKTOP_NOTIFY_BACKEND == "win10toast":
                _Win10Toast().show_toast(title, message, duration=10, threaded=True)
            log(f"Desktop notification sent: {title}")
        except Exception as e:
            log(f"Desktop notification failed: {e}", "warning")

    # ---------------- Message box popup ----------------

    def show_messagebox(self, title: str, message: str, icon: str = "info"):
        """Pop up a Windows message box. Runs in a background daemon thread
        so it never blocks cleanup/shutdown - the popup just sits on screen
        (e.g. on the task-scheduler's desktop) until someone clicks OK.
        """
        if not Config.MESSAGEBOX_ENABLED:
            return
        if platform.system() != "Windows":
            log("Message box skipped: not running on Windows", "warning")
            return
        try:
            import ctypes
            icon_flags = {"info": 0x40, "warning": 0x30, "error": 0x10}
            flags = icon_flags.get(icon, 0x40) | 0x1000  # | MB_SYSTEMMODAL so it's on top

            def _show():
                try:
                    ctypes.windll.user32.MessageBoxW(0, message, title, flags)
                except Exception as e:
                    log(f"Message box failed: {e}", "warning")

            threading.Thread(target=_show, daemon=True).start()
            log(f"Message box shown: {title}")
        except Exception as e:
            log(f"Message box failed: {e}", "warning")

    # ---------------- Email ----------------

    def send_email(self, subject: str, body: str, attachments: Optional[List[Path]] = None):
        if not Config.EMAIL_ENABLED:
            return
        if not Config.EMAIL_USERNAME or not Config.EMAIL_TO:
            log("Email notification skipped: EMAIL_USERNAME / EMAIL_TO not configured", "warning")
            return

        try:
            msg = MIMEMultipart()
            msg["From"] = Config.EMAIL_FROM or Config.EMAIL_USERNAME
            msg["To"] = ", ".join(Config.EMAIL_TO)
            msg["Subject"] = subject
            msg.attach(MIMEText(body, "plain"))

            for path in (attachments or []):
                try:
                    path = Path(path)
                    if path.exists():
                        with open(path, "rb") as f:
                            part = MIMEApplication(f.read(), Name=path.name)
                        part["Content-Disposition"] = f'attachment; filename="{path.name}"'
                        msg.attach(part)
                except Exception as e:
                    log(f"Could not attach {path}: {e}", "warning")

            context = ssl.create_default_context()
            with smtplib.SMTP(Config.EMAIL_SMTP_HOST, Config.EMAIL_SMTP_PORT) as server:
                server.starttls(context=context)
                server.login(Config.EMAIL_USERNAME, Config.EMAIL_PASSWORD)
                server.sendmail(msg["From"], Config.EMAIL_TO, msg.as_string())

            log(f"Email notification sent: {subject}")
        except Exception as e:
            log(f"Email notification failed: {e}", "warning")

    # ---------------- Combined helpers ----------------

    def notify_start(self):
        self.send_desktop("AR Automation Started", "Login successful, processing reports...")
        if Config.EMAIL_NOTIFY_ON_START:
            self.send_email("AR Automation Started", "The AR automation run has started.")

    def notify_run_complete(self, summary: Dict[str, Any], attachments: Optional[List[Path]] = None):
        failed = len(summary.get("failed", []))
        partial = len(summary.get("partial", []))
        completed = len(summary.get("completed", []))

        title = "AR Automation Completed" if failed == 0 else "AR Automation Completed (with issues)"
        message = f"Completed: {completed} | Partial: {partial} | Failed: {failed}"
        self.send_desktop(title, message)
        self.show_messagebox(title, message, icon="info" if failed == 0 else "warning")

        should_email = (
            (failed > 0 or partial > 0) and Config.EMAIL_NOTIFY_ON_FAILURE
        ) or (failed == 0 and partial == 0 and Config.EMAIL_NOTIFY_ON_SUCCESS)

        if should_email:
            body_lines = [
                f"AR Automation run finished at {datetime.now():%Y-%m-%d %H:%M:%S}", "",
                f"Completed: {completed}",
                f"Partial:   {partial}",
                f"Failed:    {failed}", "",
            ]
            if summary.get("failed"):
                body_lines.append("Failed practices:")
                body_lines += [f"  - {p}" for p in summary["failed"]]
                body_lines.append("")
            if summary.get("partial"):
                body_lines.append("Partially completed practices:")
                body_lines += [f"  - {p}" for p in summary["partial"]]

            self.send_email(title, "\n".join(body_lines), attachments=attachments)

    def notify_fatal_error(self, error: Exception, screenshot: Optional[Path] = None):
        title = "AR Automation FAILED"
        message = str(error)[:200]
        self.send_desktop(title, message)
        self.show_messagebox(title, message, icon="error")
        if Config.EMAIL_NOTIFY_ON_FAILURE:
            body = f"AR Automation stopped with a fatal error:\n\n{error}\n\n{traceback.format_exc()}"
            self.send_email(title, body, attachments=[screenshot] if screenshot else None)


notification_manager = NotificationManager()

# ==========================================================
# CIRCUIT BREAKER
# ==========================================================

class CircuitBreaker:
    """Tracks consecutive practice failures and trips (raises) once a
    threshold is hit, so a dead session / down site doesn't burn through
    every remaining practice's retries for no benefit.

    A single success resets the consecutive-failure counter, since that
    proves the session/site is actually working.
    """

    def __init__(self, threshold: int = Config.CIRCUIT_BREAKER_THRESHOLD,
                 enabled: bool = Config.CIRCUIT_BREAKER_ENABLED):
        self.threshold = threshold
        self.enabled = enabled
        self.consecutive_failures = 0
        self.tripped = False

    def record_success(self):
        self.consecutive_failures = 0

    def record_failure(self, practice_name: str = ""):
        if not self.enabled:
            return
        self.consecutive_failures += 1
        log(f"Circuit breaker: {self.consecutive_failures}/{self.threshold} consecutive failures", "warning")

        if self.consecutive_failures >= self.threshold:
            self.tripped = True
            raise CircuitBreakerTripped(
                f"{self.consecutive_failures} practices failed in a row "
                f"(last: '{practice_name}'). Stopping run to avoid wasting further retries - "
                f"check login/session/site status."
            )

    def reset(self):
        self.consecutive_failures = 0
        self.tripped = False


# ==========================================================
# DECORATORS
# ==========================================================

def retry_on_exception(max_retries: int = Config.MAX_RETRIES, 
                       delay: int = Config.RETRY_DELAY,
                       exceptions: Tuple = (Exception,)):
    """Decorator for automatic retry with exponential backoff."""

    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            last_error = None
            for attempt in range(1, max_retries + 1):
                try:
                    log(f"Attempt {attempt}/{max_retries} for {func.__name__}")
                    return func(*args, **kwargs)
                except exceptions as e:
                    last_error = e
                    log(f"Attempt {attempt} failed: {str(e)[:100]}", "warning")
                    if attempt < max_retries:
                        sleep_time = delay * (2 ** (attempt - 1))  # Exponential backoff
                        log(f"Retrying in {sleep_time}s...")
                        time.sleep(sleep_time)
            raise last_error
        return wrapper
    return decorator


def safe_operation(fallback_value: Any = None):
    """Decorator for safe execution with fallback."""

    def decorator(func: Callable) -> Callable:
        @wraps(func)
        def wrapper(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as e:
                log(f"Safe operation failed in {func.__name__}: {e}", "warning")
                return fallback_value
        return wrapper
    return decorator

# ==========================================================
# DOWNLOAD MANAGEMENT
# ==========================================================

class DownloadManager:
    """Manages file downloads with monitoring and cleanup."""

    def __init__(self, download_dir: Path):
        self.download_dir = download_dir
        self._lock = threading.Lock()

    @staticmethod
    def _strip_duplicate_suffix(filename: str) -> str:
        """Normalize browser-renamed files like 'Report (1).xlsx' -> 'Report.xlsx'."""
        path = Path(filename)
        clean_stem = re.sub(r"\s*\(\d+\)$", "", path.stem).strip()
        return f"{clean_stem}{path.suffix}"

    def clean_download_folder(self):
        """Remove old downloaded files."""
        extensions = (".xlsx", ".xls", ".pdf", ".crdownload", ".tmp")
        removed = 0

        for file in self.download_dir.iterdir():
            if file.suffix.lower() in extensions:
                try:
                    file.unlink()
                    removed += 1
                except Exception as e:
                    log(f"Could not remove {file}: {e}", "warning")

        log(f"Cleaned {removed} files from download folder")

    def get_existing_files(self) -> set:
        """Get current set of downloaded files."""
        return set(
            f.name for f in self.download_dir.iterdir()
            if f.suffix.lower() in (".xlsx", ".xls", ".pdf")
        )

    def wait_for_download(self, existing_files: set, timeout: int = Config.DOWNLOAD_TIMEOUT) -> str:
        """Wait for new file download with progress monitoring."""

        end_time = time.time() + timeout
        last_progress = time.time()

        while time.time() < end_time:
            # Check for partial downloads
            partial_files = list(self.download_dir.glob("*.crdownload"))

            if partial_files:
                if time.time() - last_progress > 10:
                    log(f"Download in progress... {partial_files[0].name}")
                    last_progress = time.time()
                time.sleep(1)
                continue

            # Check for completed downloads
            current_files = self.get_existing_files()
            new_files = current_files - existing_files

            if new_files:
                time.sleep(2)  # Let Chrome finish writing
                newest = max(
                    new_files,
                    key=lambda f: (self.download_dir / f).stat().st_ctime
                )
                log(f"Download completed: {newest}")
                return newest

            time.sleep(1)

        raise TimeoutException(f"Download timeout after {timeout}s")

    def move_report(self, filename: str, practice: str, report_name: str, keep_original_name: bool = False) -> Path:
        """Move downloaded file to organized directory."""

        source = self.download_dir / filename

        if not source.exists():
            raise FileNotFoundError(f"Downloaded file missing: {filename}")

        # Create practice folder
        practice_folder = self.download_dir / practice
        practice_folder.mkdir(exist_ok=True)

        # Determine destination filename
        if keep_original_name:
            # Keep original name, but normalize browser duplicate suffixes
            normalized_name = self._strip_duplicate_suffix(filename)
            destination = practice_folder / normalized_name
        else:
            # Use custom report name
            extension = source.suffix
            destination = practice_folder / f"{report_name}{extension}"

        # Remove existing destination with retry so we overwrite the
        # canonical filename instead of keeping browser-generated duplicates.
        for attempt in range(3):
            if not destination.exists():
                break
            try:
                destination.unlink()
                log(f"Removed existing: {destination.name}")
                break
            except PermissionError:
                if attempt == 2:
                    raise
                time.sleep(1)

        # Move file
        shutil.move(str(source), str(destination))
        log(f"Moved to: {destination}")

        return destination


# ==========================================================
# EXCEL MANAGEMENT
# ==========================================================

class ExcelManager:
    """Manages Excel file operations with validation."""

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self._workbook: Optional[openpyxl.Workbook] = None
        self._sheets: dict = {}

    def load(self) -> 'ExcelManager':
        """Load workbook."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        self._workbook = openpyxl.load_workbook(self.file_path)
        log(f"Excel loaded: {self.file_path.name}")
        return self

    def get_sheet(self, name: str):
        """Get worksheet by name."""
        if name not in self._workbook.sheetnames:
            raise ValueError(f"Sheet '{name}' not found. Available: {self._workbook.sheetnames}")
        return self._workbook[name]

    def read_login(self) -> Credentials:
        """Read login credentials from Login sheet."""
        df = pd.read_excel(self.file_path, sheet_name="Login", dtype=str).fillna("")

        if df.empty:
            raise ValueError("Login sheet is empty")

        row = df.iloc[0]

        return Credentials(
            username=row["Username"].strip(),
            password=row["Password"].strip(),
            security_key=row["Security Key"].strip()
        )

    def get_report_code(self, sheet_name: str, default: str) -> str:
        """Read the report search code from the sheet's own Type column
        (column A, first data row) instead of hardcoding "AGE24"/"FIN18"
        in the code. Falls back to `default` if the sheet/column is empty."""
        try:
            ws = self.get_sheet(sheet_name)
            for row in range(2, ws.max_row + 1):
                val = str(ws.cell(row, 1).value or "").strip()
                if val:
                    log(f"Report search code for '{sheet_name}' sheet taken from Excel Type column: '{val}'")
                    return val.upper()
        except Exception as e:
            log(f"Could not read Type column from '{sheet_name}' sheet ({e}); using default '{default}'", "warning")

        log(f"No Type value found in '{sheet_name}' sheet - using default '{default}'", "warning")
        return default

    def read_age24_practices(self) -> List[AGE24Practice]:
        """Read AGE24 practices from the 'AGE24' sheet.
        Columns: A=Type, B=Practice, C=Date, D=Status."""
        ws = self.get_sheet("AGE24")
        practices = []

        for row in range(2, ws.max_row + 1):
            row_type = str(ws.cell(row, 1).value or "").strip()
            practice = str(ws.cell(row, 2).value or "").strip()
            service_date = ws.cell(row, 3).value
            status = str(ws.cell(row, 4).value or "").strip().lower()

            if not practice:
                continue

            # The Type column should say AGE24 on this sheet - if a row
            # was mistakenly filled with a different type, skip it rather
            # than silently mis-processing it.
            if row_type and row_type.upper() != "AGE24":
                log(f"AGE24 row {row}: Type='{row_type}' (not AGE24) - skipping", "warning")
                continue

            if isinstance(service_date, datetime):
                service_date = service_date.strftime("%m/%d/%Y")
            else:
                service_date = str(service_date).strip()

            if Config.FORCE_REPROCESS:
                invoice_done, service_done, status_final = False, False, ""
            else:
                invoice_done = "invoice" in status or "downloaded" in status
                service_done = "service" in status or "downloaded" in status
                status_final = status

            practices.append(AGE24Practice(
                practice_type=row_type or "AGE24",
                practice_name=practice,
                service_date=service_date,
                status=status_final,
                row_number=row,
                invoice_downloaded=invoice_done,
                service_downloaded=service_done
            ))

        log(f"Loaded {len(practices)} AGE24 practices")
        return practices

    def read_fin18_practices(self) -> List[FIN18Practice]:
        """Read FIN18 practices from the 'FIN18' sheet.
        Columns: A=Type, B=Practice, C=From Date, D=To Date, E=Status."""
        ws = self.get_sheet("FIN18")
        practices = []

        def format_date(date_val):
            if isinstance(date_val, datetime):
                return date_val.strftime("%m/%d/%Y")
            elif date_val and str(date_val).strip():
                try:
                    dt = datetime.strptime(str(date_val).strip(), "%m/%d/%Y")
                    return dt.strftime("%m/%d/%Y")
                except:
                    return str(date_val).strip()
            return ""

        for row in range(2, ws.max_row + 1):
            row_type = str(ws.cell(row, 1).value or "").strip()
            practice = str(ws.cell(row, 2).value or "").strip()
            from_date = ws.cell(row, 3).value
            to_date = ws.cell(row, 4).value
            status = str(ws.cell(row, 5).value or "").strip().lower()

            if not practice:
                continue

            if row_type and row_type.upper() != "FIN18":
                log(f"FIN18 row {row}: Type='{row_type}' (not FIN18) - skipping", "warning")
                continue

            from_date_str = format_date(from_date)
            to_date_str = format_date(to_date)

            if not from_date_str or not to_date_str:
                log(f"FIN18 row {row} ({practice}): missing From/To date - skipping", "warning")
                continue

            status_final = "" if Config.FORCE_REPROCESS else status

            practices.append(FIN18Practice(
                practice_type=row_type or "FIN18",
                practice_name=practice,
                from_date=from_date_str,
                to_date=to_date_str,
                status=status_final,
                row_number=row
            ))

        log(f"Loaded {len(practices)} FIN18 practices")
        return practices

    def update_status(self, sheet_name: str, row: int, column: int, status: Status):
        """Update cell status with color coding."""

        ws = self.get_sheet(sheet_name)
        cell = ws.cell(row, column)
        cell.value = status.value

        # Color mapping
        colors = {
            Status.DOWNLOADED: "92D050",
            Status.FAILED: "FF0000",
            Status.RETRY: "FFFF00",
            Status.SKIPPED: "D9D9D9",
            Status.PENDING: "FFFFFF",
            Status.PARTIAL: "FFC000"  # Orange for partial
        }

        color = colors.get(status, "FFFFFF")
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

        # Auto-save
        self.save()
        log(f"Status updated: Row {row} = {status.value}")

    def save(self):
        """Save workbook."""
        if self._workbook:
            self._workbook.save(self.file_path)

    def close(self):
        """Close workbook."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None


# ==========================================================
# CHROME DRIVER MANAGER
# ==========================================================

class ChromeDriverManager:
    """Manages Chrome WebDriver lifecycle."""

    def __init__(self):
        self.driver: Optional[webdriver.Chrome] = None
        self._download_manager = DownloadManager(Config.DOWNLOAD_DIR)
        self.wait_utils: Optional['WaitUtils'] = None

    def create_driver(self) -> webdriver.Chrome:
        """Create optimized Chrome driver."""

        chrome_options = Options()
        chrome_options.page_load_strategy = Config.PAGE_LOAD_STRATEGY

        # Download preferences
        prefs = {
            "download.default_directory": str(Config.DOWNLOAD_DIR),
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled": True,
            "profile.default_content_setting_values.automatic_downloads": 1,
            "download.extensions_to_open": "applications/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        chrome_options.add_experimental_option("prefs", prefs)

        # Performance arguments
        if Config.HEADLESS:
            chrome_options.add_argument("--headless=new")
            chrome_options.add_argument(f"--window-size={Config.WINDOW_SIZE}")
        else:
            chrome_options.add_argument("--start-maximized")

        chrome_options.add_argument("--disable-notifications")
        chrome_options.add_argument("--disable-popup-blocking")
        chrome_options.add_argument("--disable-infobars")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-extensions")
        chrome_options.add_argument("--remote-allow-origins=*")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        if Config.DISABLE_IMAGES:
            chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument("--js-flags=--max-old-space-size=2048")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation", "enable-logging"])
        chrome_options.add_experimental_option("useAutomationExtension", False)

        # Create driver
        # Strategy:
        #   1) Try Selenium's built-in "Selenium Manager" (Selenium >= 4.6).
        #      It auto-detects the installed Chrome version (works fine with
        #      modern Chrome-for-Testing releases like 150.x) and downloads a
        #      matching chromedriver automatically - no extra package needed.
        #   2) If that fails for any reason, fall back to webdriver_manager
        #      (if installed).
        #   3) If that also fails, fall back to a local chromedriver.exe
        #      placed next to this script (manual download fallback).
        last_error = None
        driver_created = False

        # --- Attempt 1: Selenium Manager (recommended, no service path needed) ---
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
            driver_created = True
            log("Chrome driver created via built-in Selenium Manager")
        except Exception as e:
            last_error = e
            log(f"Selenium Manager driver creation failed: {e}", "warning")

        # --- Attempt 2: webdriver_manager fallback ---
        if not driver_created and _WDM_AVAILABLE:
            try:
                service = Service(WDM_ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=chrome_options)
                driver_created = True
                log("Chrome driver created via webdriver_manager fallback")
            except Exception as e:
                last_error = e
                log(f"webdriver_manager driver creation failed: {e}", "warning")

        # --- Attempt 3: local chromedriver.exe next to this script ---
        if not driver_created:
            local_driver_path = Path(__file__).resolve().parent / "chromedriver.exe"
            if local_driver_path.exists():
                try:
                    service = Service(str(local_driver_path))
                    self.driver = webdriver.Chrome(service=service, options=chrome_options)
                    driver_created = True
                    log(f"Chrome driver created via local chromedriver at {local_driver_path}")
                except Exception as e:
                    last_error = e
                    log(f"Local chromedriver.exe creation failed: {e}", "warning")

        if not driver_created:
            log(
                "Could not start Chrome driver. Make sure Google Chrome is "
                "installed and up to date, and that this machine has internet "
                "access (Selenium Manager needs to download a matching driver "
                "on first run).",
                "error"
            )
            raise last_error if last_error else RuntimeError("Chrome driver could not be created")

        try:
            self.driver.set_page_load_timeout(Config.PAGE_LOAD_TIMEOUT)
            self.driver.implicitly_wait(Config.IMPLICIT_WAIT)

            if not Config.HEADLESS:
                self.driver.maximize_window()

            self._configure_download_behavior()
            self._post_start_browser_stabilization()

            log(f"Chrome driver created successfully (headless={Config.HEADLESS})")

            # Cleanup downloads
            self._download_manager.clean_download_folder()

            return self.driver

        except Exception as e:
            log(f"Failed to create Chrome driver: {e}", "error")
            raise

    def _configure_download_behavior(self):
        """Ensure downloads work reliably in headless Chrome."""
        if not self.driver:
            return

        try:
            self.driver.execute_cdp_cmd(
                "Page.setDownloadBehavior",
                {
                    "behavior": "allow",
                    "downloadPath": str(Config.DOWNLOAD_DIR)
                }
            )
            log("Chrome download behavior configured")
        except Exception as e:
            log(f"Could not configure download behavior: {e}", "warning")

    def _post_start_browser_stabilization(self):
        """Extra stabilization for headless mode startup."""
        if not self.driver:
            return

        try:
            self.driver.execute_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )
        except Exception:
            pass

        if Config.HEADLESS and Config.HEADLESS_STARTUP_DELAY > 0:
            time.sleep(Config.HEADLESS_STARTUP_DELAY)

    def quit(self):
        """Safely quit driver."""
        if self.driver:
            try:
                self.close_extra_windows()
                self.driver.quit()
                log("Chrome driver quit successfully")
            except Exception as e:
                log(f"Error quitting driver: {e}", "warning")
            finally:
                self.driver = None

    def close_extra_windows(self):
        """Close all extra browser windows."""
        try:
            if not self.driver or len(self.driver.window_handles) <= 1:
                return

            main = self.driver.window_handles[0]

            for handle in self.driver.window_handles[1:]:
                try:
                    self.driver.switch_to.window(handle)
                    self.driver.close()
                except Exception:
                    pass

            self.driver.switch_to.window(main)
            log("Extra windows closed")
        except Exception as e:
            log(f"close_extra_windows skipped - browser session unavailable: {e}", "warning")

    def reset_frame(self):
        """Reset to default content."""
        try:
            self.driver.switch_to.default_content()
            time.sleep(0.5)
        except Exception:
            pass

    def save_screenshot(self, name: str) -> Optional[Path]:
        """Save screenshot with timestamp. Returns the file path, or None on failure."""
        try:
            Config.SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
            filename = Config.SCREENSHOT_DIR / f"{name}_{datetime.now():%Y%m%d_%H%M%S}.png"
            self.driver.save_screenshot(str(filename))
            log(f"Screenshot saved: {filename.name}")
            return filename
        except Exception as e:
            log(f"Screenshot failed: {e}", "warning")
            return None

    def is_session_alive(self) -> bool:
        """Check if session is still valid."""
        try:
            self.driver.find_element(By.XPATH, "//*[@id='tdMenuBarItemPatient']/a")
            return True
        except:
            return False

    def hard_reset(self):
        """Hard reset: close all popups, reset frames, refresh if needed."""
        log("Performing hard reset...")
        try:
            self.close_extra_windows()
            self.reset_frame()
            # Check if we're still on a valid page
            current_url = self.driver.current_url
            if "practicevelocity" not in current_url:
                log("Page seems invalid, navigating back to base URL", "warning")
                self.driver.get(Config.URL)
                time.sleep(3)
            return True
        except Exception as e:
            log(f"Hard reset failed: {e}", "error")
            return False


# ==========================================================
# WAIT UTILITIES
# ==========================================================

class WaitUtils:
    """Advanced wait utilities."""

    def __init__(self, driver: webdriver.Chrome):
        self.driver = driver

    def wait(self, timeout: int = Config.ELEMENT_TIMEOUT) -> WebDriverWait:
        """Get WebDriverWait instance."""
        return WebDriverWait(self.driver, timeout)

    def safe_click(self, locator: Tuple, timeout: int = Config.ELEMENT_TIMEOUT, use_js: bool = False):
        """Safe click with JavaScript fallback and multiple strategies."""

        # Strategy 1: Standard clickable wait
        try:
            element = self.wait(timeout).until(
                EC.element_to_be_clickable(locator)
            )
            element.click()
            return element
        except (ElementClickInterceptedException, ElementNotInteractableException) as e:
            log(f"Standard click failed, trying JS click: {e}", "warning")
            use_js = True
        except Exception as e:
            log(f"Click error: {e}", "warning")
            use_js = True

        # Strategy 2: JavaScript click
        if use_js:
            try:
                element = self.wait(timeout).until(
                    EC.presence_of_element_located(locator)
                )
                self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                time.sleep(0.5)
                self.driver.execute_script("arguments[0].click();", element)
                log("JS click succeeded")
                return element
            except Exception as e2:
                log(f"JS click also failed: {e2}", "error")
                raise

    def safe_type(self, locator: Tuple, value: str, timeout: int = Config.ELEMENT_TIMEOUT):
        """Safe text input with clear and focus."""
        textbox = self.wait(timeout).until(
            EC.presence_of_element_located(locator)
        )

        # Focus and clear using multiple methods
        try:
            textbox.click()
        except:
            self.driver.execute_script("arguments[0].focus();", textbox)

        # Select all and clear
        textbox.send_keys("\ue01a")  # Ctrl+A
        time.sleep(0.2)
        textbox.send_keys("\ue017")  # Delete key
        time.sleep(0.2)

        # Type new value
        textbox.send_keys(str(value))

        # Verify
        actual = textbox.get_attribute("value") or ""
        if str(value) not in actual:
            log(f"Type verification failed, trying JS set value", "warning")
            self.driver.execute_script(f"arguments[0].value = '{value}';", textbox)
            self.driver.execute_script("arguments[0].dispatchEvent(new Event('change'));", textbox)

        return textbox

    def wait_for_popup(self, timeout: int = Config.POPUP_TIMEOUT) -> Tuple[str, str]:
        """Wait for popup window and switch to it."""
        self.wait(timeout).until(lambda d: len(d.window_handles) > 1)

        main = self.driver.current_window_handle
        popup = [h for h in self.driver.window_handles if h != main][0]

        self.driver.switch_to.window(popup)
        self.wait().until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

        log("Popup loaded")
        return main, popup

    def switch_frame(self, locator: Tuple, timeout: int = Config.ELEMENT_TIMEOUT):
        """Switch to frame with retry."""
        self.wait(timeout).until(
            EC.frame_to_be_available_and_switch_to_it(locator)
        )

    def wait_for_element_stable(self, locator: Tuple, timeout: int = 10):
        """Wait for element to be present and stable (not moving)."""
        element = self.wait(timeout).until(
            EC.presence_of_element_located(locator)
        )
        # Wait a bit for any animations
        time.sleep(0.5)
        return element


# ==========================================================
# LOGIN MODULE
# ==========================================================

class LoginModule:
    """Handles authentication with TOTP."""

    def __init__(self, driver: webdriver.Chrome, wait_utils: WaitUtils):
        self.driver = driver
        self.wait = wait_utils

    @retry_on_exception(max_retries=2, delay=2)
    def login(self, credentials: Credentials) -> bool:
        """Perform full login sequence."""

        self.driver.get(Config.URL)

        # Generate OTP
        otp = pyotp.TOTP(
            credentials.security_key.replace(" ", "").upper()
        ).now()

        log(f"Logging in as: {credentials.username}")
        log(f"Generated OTP: {otp}")

        # Step 1: Username - explicitly clear first to avoid 0 prefix
        username_field = self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "txtLogin"))
        )
        username_field.clear()
        time.sleep(0.5)
        username_field.send_keys(credentials.username)
        time.sleep(0.3)
        # Verify no extra characters
        actual_user = username_field.get_attribute("value") or ""
        if actual_user != credentials.username:
            log(f"Username field mismatch! Expected: '{credentials.username}', Got: '{actual_user}'", "warning")
            self.driver.execute_script("arguments[0].value = arguments[1];", username_field, credentials.username)
        log(f"Username entered: '{actual_user}'")
        self.wait.safe_click((By.ID, "btnNext"))

        # Step 2: Continue
        self.wait.safe_click((By.XPATH, "//*[@id='form20']/div[2]/input"))

        # Step 3: Password - explicitly clear first to avoid 0 prefix
        password_field = self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "input54"))
        )
        password_field.clear()
        time.sleep(0.5)
        password_field.send_keys(credentials.password)
        time.sleep(0.3)
        # Verify no extra characters (password is masked, so check length)
        actual_pass = password_field.get_attribute("value") or ""
        if len(actual_pass) != len(credentials.password):
            log(f"Password field length mismatch! Expected: {len(credentials.password)}, Got: {len(actual_pass)}", "warning")
            password_field.clear()
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].value = arguments[1];", password_field, credentials.password)
            self.driver.execute_script("arguments[0].dispatchEvent(new Event('input'));", password_field)
        log(f"Password entered (length: {len(actual_pass)})")

        self.wait.safe_click((By.XPATH, "//*[@id='form46']/div[2]/input"))

        # Step 4: OTP - explicitly clear first to avoid 0 prefix
        otp_field = self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "input80"))
        )
        otp_field.clear()
        time.sleep(0.5)
        otp_field.send_keys(otp)
        time.sleep(0.3)
        # Verify OTP
        actual_otp = otp_field.get_attribute("value") or ""
        if actual_otp != otp:
            log(f"OTP field mismatch! Expected: '{otp}', Got: '{actual_otp}'", "warning")
            otp_field.clear()
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].value = arguments[1];", otp_field, otp)
            self.driver.execute_script("arguments[0].dispatchEvent(new Event('input'));", otp_field)
        log(f"OTP entered: '{actual_otp}'")

        self.wait.safe_click((By.XPATH, "//*[@id='form72']/div[2]/input"))

        # Verify login
        self.wait.wait(Config.POPUP_TIMEOUT).until(
            EC.presence_of_element_located(
                (By.XPATH, "//*[@id='tdMenuBarItemPatient']/a")
            )
        )

        log("Login successful")
        return True

    def ensure_login(self, credentials: Credentials) -> bool:
        """Check session and re-login if needed."""
        try:
            if self.driver.find_element(By.XPATH, "//*[@id='tdMenuBarItemPatient']/a"):
                return True
        except:
            pass

        log("Session expired - re-authenticating")
        return self.login(credentials)


# ==========================================================
# NAVIGATION MODULE
# ==========================================================

class NavigationModule:
    """Handles page navigation and frame switching."""

    def __init__(self, driver: webdriver.Chrome, wait_utils: WaitUtils):
        self.driver = driver
        self.wait = wait_utils

    def open_administration(self):
        """Open Administration > Reports."""
        self.driver.switch_to.default_content()
        self.wait.safe_click((By.XPATH, "//*[@id='tdMenuBarItemAdministration']/a"))
        time.sleep(1)
        self.wait.safe_click((By.XPATH, "//*[@id='menu_Administration_Reports']"))
        log("Administration > Reports opened")

    def switch_report_frame(self):
        """Switch to report main frame."""
        self.driver.switch_to.default_content()
        self.wait.switch_frame((By.ID, "reportMainWindow"))

    def switch_nav_frame(self):
        """Switch to navigation frame."""
        self.switch_report_frame()
        self.wait.switch_frame((By.NAME, "NavFrame"))

    def switch_main_frame(self):
        """Switch to main content frame."""
        self.switch_report_frame()
        self.wait.switch_frame((By.NAME, "PVRC_MainStage"))

    def search_report(self, report_code: str):
        """Search for report by code - BULLETPROOF no 0 prefix."""
        from selenium.webdriver.common.keys import Keys

        # Full frame reset before search
        self.driver.switch_to.default_content()
        time.sleep(1)

        self.switch_nav_frame()

        # Find search box directly
        search_box = self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "userSearch"))
        )

        # Click focus
        search_box.click()
        time.sleep(0.3)

        # Select all and delete
        search_box.send_keys(Keys.CONTROL, "a")
        time.sleep(0.2)
        search_box.send_keys(Keys.DELETE)
        time.sleep(0.3)

        # Verify empty
        current = search_box.get_attribute("value") or ""
        if current != "":
            self.driver.execute_script("arguments[0].value = '';", search_box)
            time.sleep(0.2)

        # JS set value directly - NO 0 prefix
        self.driver.execute_script("arguments[0].value = arguments[1];", search_box, report_code)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", search_box)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", search_box)
        time.sleep(0.3)

        # Verify
        actual = search_box.get_attribute("value") or ""
        if actual != report_code:
            log(f"Search box mismatch! Expected '{report_code}', got '{actual}'", "warning")

        log(f"Report search entered: '{actual}'")

        # Click search button
        self.wait.safe_click((By.ID, "dosearch"))
        log(f"Report search submitted: {report_code}")

        self.switch_main_frame()

        # After searching, Experity shows a RESULTS LIST first (e.g. "All
        # Results" with report titles like "Non-Month End Aging"). We need
        # to click into the matching result before the practice dropdown
        # ("subpracticeselect") shows up. Some searches may load the report
        # directly, so check for the dropdown briefly before assuming a
        # results page is in the way.
        try:
            self.wait.wait(6).until(
                EC.presence_of_element_located((By.ID, "subpracticeselect"))
            )
            log(f"{report_code} loaded directly (no results page)")
            return
        except TimeoutException:
            log(f"No direct report load for {report_code} - looking for results list", "info")

        self._click_search_result(report_code)

        self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "subpracticeselect"))
        )
        log(f"{report_code} loaded")

    def _click_search_result(self, report_code: str):
        """Click the matching report link on the search-results page."""

        link_text = Config.REPORT_RESULT_LINK_TEXT.get(report_code)

        if link_text:
            try:
                link = self.wait.wait(10).until(
                    EC.element_to_be_clickable((By.XPATH, f"//a[normalize-space()='{link_text}']"))
                )
                link.click()
                log(f"Clicked search result: '{link_text}'")
                return
            except Exception as e:
                log(f"Configured result link '{link_text}' not found/clickable ({e}); "
                    f"falling back to first result", "warning")

        # Fallback: click the first link that appears in the results list.
        # The search term is specific (e.g. "AGE24"), so the top result is
        # normally the right one.
        try:
            first_link = self.wait.wait(10).until(
                EC.element_to_be_clickable((By.XPATH, "(//div[@id='AllResults']//a | //*[@id='ResultTable']//a)[1]"))
            )
            log(f"Clicking first search result: '{first_link.text.strip()}'")
            first_link.click()
        except Exception:
            # Last-resort generic fallback: any bold/report-style link below
            # the "Search for '<code>'" heading.
            first_link = self.wait.wait(10).until(
                EC.element_to_be_clickable(
                    (By.XPATH, f"//*[contains(text(), \"Search for\")]/following::a[1]")
                )
            )
            log(f"Clicking first search result (generic fallback): '{first_link.text.strip()}'")
            first_link.click()


# ==========================================================
# REPORT OPERATIONS
# ==========================================================

class ReportOperations:
    """Handles report configuration and execution."""

    def __init__(self, driver: webdriver.Chrome, wait_utils: WaitUtils):
        self.driver = driver
        self.wait = wait_utils

    def select_practice(self, practice: str):
        """Select practice - handles BOTH the classic single dropdown
        (id='subpracticeselect', regular logins) AND the multi-clinic
        checkbox panel ("Clinics" with Check All / Uncheck All buttons,
        seen on Super Practice logins)."""

        # Try the classic dropdown first
        try:
            dropdown = self.wait.wait(6).until(
                EC.presence_of_element_located((By.ID, "subpracticeselect"))
            )
            time.sleep(1)
            select = Select(dropdown)

            try:
                select.select_by_visible_text(practice)
            except NoSuchElementException:
                options = [opt.text.strip() for opt in select.options]
                matching = [opt for opt in options if practice.lower() in opt.lower()]
                if matching:
                    select.select_by_visible_text(matching[0])
                    log(f"Practice matched via partial: {matching[0]}")
                else:
                    raise Exception(f"Practice '{practice}' not found in dropdown. Available: {options[:10]}...")

            log(f"Practice selected (dropdown): {practice}")
            time.sleep(1)
            return

        except TimeoutException:
            log("No practice dropdown found - trying Clinics checkbox panel (Super Practice view)", "info")

        # Super Practice view: "Clinics" checkbox panel
        self._select_clinic_checkbox(practice)

    def _find_clinic_checkbox(self, practice: str):
        """Locate a Clinics-panel checkbox for the given practice code/name.
        Returns the WebElement, or None if not found (never raises)."""
        strategies = [
            (By.XPATH, f"//input[@type='checkbox'][@value='{practice}']"),
            (By.XPATH, f"//input[@type='checkbox'][@id='{practice}']"),
            (By.XPATH, f"//input[@type='checkbox'][contains(@id, '{practice}')]"),
            (By.XPATH, f"//input[@type='checkbox'][following-sibling::text()[1][normalize-space()='{practice}']]"),
            (By.XPATH, f"//td[normalize-space(text())='{practice}']/preceding-sibling::td[1]//input[@type='checkbox']"),
            (By.XPATH, f"//td[normalize-space(text())='{practice}']//input[@type='checkbox']"),
            (By.XPATH, f"//label[normalize-space(text())='{practice}']/input[@type='checkbox']"),
            (By.XPATH, f"//label[normalize-space(text())='{practice}']/preceding-sibling::input[@type='checkbox'][1]"),
            (By.XPATH, f"//*[normalize-space(text())='{practice}']/preceding::input[@type='checkbox'][1]"),
        ]
        for by, locator in strategies:
            try:
                el = self.driver.find_element(by, locator)
                return el
            except NoSuchElementException:
                continue
        return None

    def _select_clinic_checkbox(self, practice: str):
        """Select a single clinic in the checkbox-based 'Clinics' panel:
        clear everything via 'Uncheck All', then check only the target."""

        # The FIRST "Uncheck All" on the page belongs to the Clinics panel
        # (it appears above the Financial Class panel's own Check/Uncheck All).
        try:
            uncheck_all = self.wait.wait(10).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "(//input[@value='Uncheck All'] | //button[normalize-space()='Uncheck All'] "
                    "| //a[normalize-space()='Uncheck All'])[1]"
                ))
            )
            uncheck_all.click()
            time.sleep(0.5)
            log("Cleared all clinics (Uncheck All)")
        except Exception as e:
            log(f"Could not click Clinics 'Uncheck All' (continuing anyway): {e}", "warning")

        # Find and check the target clinic's checkbox
        checkbox = None
        try:
            checkbox = self.wait.wait(10).until(lambda d: self._find_clinic_checkbox(practice))
        except TimeoutException:
            pass

        if checkbox is None:
            raise Exception(
                f"Could not find a Clinics checkbox for practice '{practice}'. "
                f"The checkbox HTML may differ from what's expected - inspect "
                f"a checkbox (F12 > right-click > Inspect) and share it so the "
                f"selector list in _find_clinic_checkbox() can be updated."
            )

        self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", checkbox)
        time.sleep(0.3)

        if not checkbox.is_selected():
            try:
                checkbox.click()
            except (ElementClickInterceptedException, ElementNotInteractableException):
                self.driver.execute_script("arguments[0].click();", checkbox)

        time.sleep(0.5)
        log(f"Practice selected (checkbox): {practice}")

    def verify_practice(self, practice: str, max_attempts: int = 20):
        """Verify practice selection with retry - handles dropdown AND checkbox modes."""
        for attempt in range(max_attempts):
            try:
                # Dropdown mode
                try:
                    dropdown = self.driver.find_element(By.ID, "subpracticeselect")
                    selected = Select(dropdown).first_selected_option.text.strip()
                    if selected == practice or practice.lower() in selected.lower():
                        log(f"Practice verified (dropdown): {selected}")
                        return True
                except NoSuchElementException:
                    pass

                # Checkbox mode
                checkbox = self._find_clinic_checkbox(practice)
                if checkbox is not None and checkbox.is_selected():
                    log(f"Practice verified (checkbox): {practice}")
                    return True

            except StaleElementReferenceException:
                log(f"Stale element on attempt {attempt+1}, retrying...", "warning")
                time.sleep(1)
                continue
            except Exception as e:
                log(f"Verify error attempt {attempt+1}: {e}", "warning")

            time.sleep(1)

        raise Exception(f"Practice not changed to: {practice} after {max_attempts} attempts")

    def open_criteria(self):
        """Open report criteria panel (older layout). On newer report pages
        the Parameters panel is already visible by default, so if the
        trigger button isn't present this is a no-op instead of a failure."""
        try:
            self.wait.safe_click((By.ID, "mainbutton1"), timeout=8)
            log("Report criteria opened")
            time.sleep(1)
        except Exception as e:
            log(f"'mainbutton1' criteria toggle not found - assuming Parameters panel "
                f"is already open on this page ({str(e)[:100]})", "info")

    def enter_service_date(self, date: str):
        """Enter service date - BULLETPROOF no 0 prefix."""
        field = self.wait.wait().until(
            EC.presence_of_element_located((By.ID, "ServiceDate"))
        )
        # Click focus
        field.click()
        time.sleep(0.3)
        # Select all and clear
        from selenium.webdriver.common.keys import Keys
        field.send_keys(Keys.CONTROL, "a")
        time.sleep(0.2)
        field.send_keys(Keys.DELETE)
        time.sleep(0.3)
        # JS set value directly
        self.driver.execute_script("arguments[0].value = arguments[1];", field, date)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", field)
        self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", field)
        time.sleep(0.3)
        # Verify
        actual = field.get_attribute("value") or ""
        if actual != date:
            log(f"ServiceDate mismatch! Expected '{date}', got '{actual}'", "warning")
        log(f"Service date: {actual}")

    def enter_date_range(self, from_date: str, to_date: str):
        """Enter date range for FIN18 - BULLETPROOF no 0 prefix.

        Some sites auto-sync the To-Date field to match From-Date (or vice
        versa) via their own onchange JS the moment one field changes. To
        guard against ending up with From==To when the Excel actually had
        two different dates, this sets both fields, then re-checks BOTH
        after both are set and re-applies any field that drifted.
        """
        from selenium.webdriver.common.keys import Keys

        def set_field(field_id: str, value: str) -> str:
            field = self.wait.wait().until(
                EC.presence_of_element_located((By.ID, field_id))
            )
            field.click()
            time.sleep(0.3)
            field.send_keys(Keys.CONTROL, "a")
            time.sleep(0.2)
            field.send_keys(Keys.DELETE)
            time.sleep(0.3)
            self.driver.execute_script("arguments[0].value = arguments[1];", field, value)
            self.driver.execute_script("arguments[0].dispatchEvent(new Event('input', {bubbles: true}));", field)
            self.driver.execute_script("arguments[0].dispatchEvent(new Event('change', {bubbles: true}));", field)
            # Blur explicitly - some sites only run their date-sync/validation
            # JS on blur, not on 'change', so skipping this can leave stale state
            self.driver.execute_script("arguments[0].blur();", field)
            time.sleep(0.3)
            return field.get_attribute("value") or ""

        # Pass 1: set From, then To
        actual_from = set_field("FromServiceDate", from_date)
        actual_to = set_field("ToServiceDate", to_date)

        # Pass 2: re-read both AFTER both are set. If setting To caused the
        # site to silently resync From (or vice versa), reapply the
        # drifted one so we end up with exactly what the Excel said.
        actual_from = self.driver.find_element(By.ID, "FromServiceDate").get_attribute("value") or ""
        actual_to = self.driver.find_element(By.ID, "ToServiceDate").get_attribute("value") or ""

        if actual_from != from_date:
            log(f"FromServiceDate drifted to '{actual_from}' after To was set - reapplying", "warning")
            actual_from = set_field("FromServiceDate", from_date)

        if actual_to != to_date:
            log(f"ToServiceDate drifted to '{actual_to}' - reapplying", "warning")
            actual_to = set_field("ToServiceDate", to_date)

        # Final check
        if actual_from != from_date:
            log(f"FromServiceDate mismatch! Expected '{from_date}', got '{actual_from}'", "warning")
        if actual_to != to_date:
            log(f"ToServiceDate mismatch! Expected '{to_date}', got '{actual_to}'", "warning")

        log(f"Date range: {actual_from} to {actual_to}")

    def select_invoice_date(self):
        """Select invoice date option."""
        invoice_xpath = '//*[@id="rightcol"]/div[1]/div[5]/div/div/div/ul/li[1]/input'
        service_xpath = '//*[@id="rightcol"]/div[1]/div[5]/div/div/div/ul/li[2]/input'

        try:
            invoice = self.wait.wait().until(
                EC.presence_of_element_located((By.XPATH, invoice_xpath))
            )
            service = self.driver.find_element(By.XPATH, service_xpath)

            # Uncheck service if checked
            if service.is_selected():
                self.driver.execute_script("arguments[0].click();", service)
                time.sleep(0.5)

            # Check invoice if not checked
            if not invoice.is_selected():
                self.driver.execute_script("arguments[0].click();", invoice)
                time.sleep(0.5)

            log("Invoice date selected")
        except Exception as e:
            log(f"Invoice date selection issue: {e}", "warning")
            # Try alternative approach
            self._toggle_date_option("invoice")

    def select_service_date(self):
        """Select service date option."""
        invoice_xpath = '//*[@id="rightcol"]/div[1]/div[5]/div/div/div/ul/li[1]/input'
        service_xpath = '//*[@id="rightcol"]/div[1]/div[5]/div/div/div/ul/li[2]/input'

        try:
            invoice = self.driver.find_element(By.XPATH, invoice_xpath)
            service = self.wait.wait().until(
                EC.presence_of_element_located((By.XPATH, service_xpath))
            )

            # Uncheck invoice if checked
            if invoice.is_selected():
                self.driver.execute_script("arguments[0].click();", invoice)
                time.sleep(0.5)

            # Check service if not checked
            if not service.is_selected():
                self.driver.execute_script("arguments[0].click();", service)
                time.sleep(0.5)

            log("Service date selected")
        except Exception as e:
            log(f"Service date selection issue: {e}", "warning")
            self._toggle_date_option("service")

    def _toggle_date_option(self, option_type: str):
        """Alternative method to toggle date option using label click."""
        try:
            if option_type == "invoice":
                label = self.driver.find_element(By.XPATH, "//label[contains(text(), 'Invoice Date') or contains(@for, 'Invoice')]")
            else:
                label = self.driver.find_element(By.XPATH, "//label[contains(text(), 'Service Date') or contains(@for, 'Service')]")

            self.driver.execute_script("arguments[0].click();", label)
            log(f"Date option toggled via label: {option_type}")
        except Exception as e2:
            log(f"Label toggle also failed: {e2}", "warning")

    def run_report(self):
        """Click run report."""
        self.wait.safe_click((By.XPATH, '//*[@id="rightcol"]/div[2]/input'))
        log("Run report clicked")
        time.sleep(2)

    def export_excel(self, download_manager: DownloadManager) -> str:
        """Export report as Excel."""

        current_url = self.driver.current_url

        if "ReportViewer.aspx" not in current_url:
            raise Exception("Not on ReportViewer page")

        export_url = current_url if "rs:Format" in current_url else current_url + "&rs:Format=EXCELOPENXML"

        log(f"Export URL prepared")

        existing_files = download_manager.get_existing_files()

        # Trigger download
        self.driver.execute_script("window.location.href = arguments[0];", export_url)
        log("Download triggered")

        time.sleep(3)

        downloaded = download_manager.wait_for_download(existing_files)
        log(f"Downloaded: {downloaded}")

        return downloaded


# ==========================================================
# DOWNLOAD MODULE (Consolidated)
# ==========================================================

class DownloadModule:
    """Consolidated download handler for all report types."""

    def __init__(self, driver_manager: ChromeDriverManager, 
                 report_ops: ReportOperations,
                 download_manager: DownloadManager):
        self.driver_manager = driver_manager
        self.report_ops = report_ops
        self.download_manager = download_manager
        self.driver = driver_manager.driver

    def download_report(self, practice: str, report_name: str, keep_original_name: bool = False) -> Path:
        """Universal download method for any report type."""

        log(f"Starting download: {report_name} for {practice}")

        # Run report
        self.report_ops.run_report()

        # Wait for popup
        main, popup = self.driver_manager.wait_utils.wait_for_popup()

        # Export
        downloaded = self.report_ops.export_excel(self.download_manager)

        # Move file
        destination = self.download_manager.move_report(downloaded, practice, report_name, keep_original_name)

        # Validate
        self._validate_excel(destination)

        # Close popup
        self._close_popup(main)

        log(f"Download completed: {destination}")
        return destination

    def _validate_excel(self, path: Path):
        """Validate Excel file integrity."""
        if not path.exists():
            raise FileNotFoundError("Downloaded file missing")

        if path.stat().st_size == 0:
            raise ValueError("Downloaded file is empty")

        try:
            openpyxl.load_workbook(path)
        except Exception:
            raise ValueError("Corrupted Excel file")

        log("Excel validation passed")

    def _close_popup(self, main_window: str):
        """Close popup and return to main window."""
        try:
            self.driver.close()
        except:
            pass

        self.driver.switch_to.window(main_window)
        self.driver.switch_to.default_content()
        time.sleep(1)
        log("Popup closed, returned to main window")


# ==========================================================
# AGE24 MODULE - FIXED: BOTH INVOICE AND SERVICE DATE
# ==========================================================

class AGE24Module:
    """Handles AGE24 report processing - BOTH Invoice AND Service Date."""

    def __init__(self, driver_manager: ChromeDriverManager,
                 navigation: NavigationModule,
                 report_ops: ReportOperations,
                 download_module: DownloadModule,
                 excel: ExcelManager,
                 auto_fix: Optional['AutoFixEngine'] = None,
                 circuit_breaker: Optional[CircuitBreaker] = None,
                 report_code: str = "AGE24"):
        self.driver_manager = driver_manager
        self.navigation = navigation
        self.report_ops = report_ops
        self.download_module = download_module
        self.excel = excel
        self.driver = driver_manager.driver
        self.wait = driver_manager.wait_utils
        self.auto_fix = auto_fix
        self.circuit_breaker = circuit_breaker
        self.last_exception: Optional[Exception] = None
        self.report_code = report_code  # search term - comes from Excel 'Type' column, not hardcoded

    def prepare(self, practice: str, service_date: str):
        """Prepare AGE24 criteria.

        IMPORTANT: always re-search "AGE24" first. After a report is
        downloaded, the browser is left on the RESULTS page (no Clinics
        panel, no practice dropdown at all) - not the parameters page.
        Since this is called separately for the Invoice Date download AND
        the Service Date download, skipping the re-search means the second
        call tries to select a practice on a page that has none of those
        elements, which fails every selector.
        """
        self.navigation.search_report(self.report_code)

        self.report_ops.select_practice(practice)
        self.report_ops.verify_practice(practice)
        self.report_ops.open_criteria()
        self.report_ops.enter_service_date(service_date)

        log("AGE24 criteria ready")

    def download_invoice_report(self, practice: AGE24Practice) -> bool:
        """Download Invoice Date report."""
        try:
            log(f"--- Downloading INVOICE DATE report for {practice.practice_name} ---")
            self.prepare(practice.practice_name, practice.service_date)
            self.report_ops.select_invoice_date()

            self.download_module.download_report(
                practice.practice_name, 
                Config.AGE24_INVOICE_SUFFIX,
                keep_original_name=False
            )
            log(f"Invoice Date report downloaded for {practice.practice_name}")
            return True
        except Exception as e:
            log(f"Invoice Date download failed: {e}", "error")
            return False

    def download_service_report(self, practice: AGE24Practice) -> bool:
        """Download Service Date report."""
        try:
            log(f"--- Downloading SERVICE DATE report for {practice.practice_name} ---")
            self.prepare(practice.practice_name, practice.service_date)
            self.report_ops.select_service_date()

            self.download_module.download_report(
                practice.practice_name, 
                Config.AGE24_SERVICE_SUFFIX,
                keep_original_name=True  # Service Date report keeps its original filename (same as FIN18)
            )
            log(f"Service Date report downloaded for {practice.practice_name}")
            return True
        except Exception as e:
            log(f"Service Date download failed: {e}", "error")
            return False

    def process_practice(self, practice: AGE24Practice) -> Tuple[bool, bool]:
        """Process single AGE24 practice - BOTH reports."""

        log("=" * 80)
        log(f"AGE24 Practice: {practice.practice_name}")
        log(f"Service Date: {practice.service_date}")
        log("=" * 80)

        invoice_success = False
        service_success = False

        try:
            # Download Invoice Date Report
            if not practice.invoice_downloaded:
                invoice_success = self.download_invoice_report(practice)
            else:
                log("Invoice report already downloaded, skipping")
                invoice_success = True

            # Download Service Date Report
            if not practice.service_downloaded:
                service_success = self.download_service_report(practice)
            else:
                log("Service report already downloaded, skipping")
                service_success = True

            # Determine overall status
            if invoice_success and service_success:
                self.excel.update_status("AGE24", practice.row_number, 4, Status.DOWNLOADED)
                state_manager.mark_completed(f"AGE24_{practice.practice_name}")
                log(f"AGE24 FULLY completed for {practice.practice_name}")
            elif invoice_success or service_success:
                self.excel.update_status("AGE24", practice.row_number, 4, Status.PARTIAL)
                state_manager.mark_partial(f"AGE24_{practice.practice_name}")
                log(f"AGE24 PARTIALLY completed for {practice.practice_name}", "warning")
            else:
                self.excel.update_status("AGE24", practice.row_number, 4, Status.FAILED)
                state_manager.mark_failed(f"AGE24_{practice.practice_name}")
                log(f"AGE24 FAILED for {practice.practice_name}", "error")

            return invoice_success, service_success

        except Exception as e:
            log(f"AGE24 failed for {practice.practice_name}: {e}", "error")
            traceback.print_exc()
            self.last_exception = e

            self.driver_manager.save_screenshot(f"AGE24_{practice.practice_name}")
            self.driver_manager.close_extra_windows()
            self.driver_manager.reset_frame()

            self.excel.update_status("AGE24", practice.row_number, 4, Status.FAILED)
            state_manager.mark_failed(f"AGE24_{practice.practice_name}")

            return False, False

    def run(self):
        """Execute AGE24 module."""
        state_manager.set_module("AGE24")

        log("=" * 80)
        log("Starting AGE24 Module (Invoice Date + Service Date)")
        log("=" * 80)

        # Navigate to AGE24
        self.navigation.search_report(self.report_code)

        # Read practices
        practices = self.excel.read_age24_practices()

        for practice in practices:
            # Skip if fully completed (unless FORCE_REPROCESS is on)
            if not Config.FORCE_REPROCESS:
                if practice.status == "downloaded":
                    log(f"Skipping (fully downloaded): {practice.practice_name}")
                    continue

                if state_manager.is_completed(f"AGE24_{practice.practice_name}"):
                    log(f"Skipping (completed in previous run): {practice.practice_name}")
                    self.excel.update_status("AGE24", practice.row_number, 4, Status.DOWNLOADED)
                    continue

            # Process with retry
            invoice_ok = False
            service_ok = False
            self.last_exception = None

            for attempt in range(1, Config.MAX_RETRIES + 1):
                log(f"Attempt {attempt}/{Config.MAX_RETRIES}")

                invoice_ok, service_ok = self.process_practice(practice)

                # If both succeeded, break
                if invoice_ok and service_ok:
                    break

                # If partial success, still retry the failed one
                if attempt < Config.MAX_RETRIES:
                    # Let the auto-fix engine diagnose whatever went wrong
                    # (stale element, blocked click, frame/session issue, etc.)
                    # before we burn another attempt on the same problem.
                    if self.auto_fix and self.last_exception:
                        self.auto_fix.diagnose_and_fix(
                            self.last_exception, context=f"AGE24_{practice.practice_name}"
                        )

                    time.sleep(Config.RETRY_DELAY * attempt)
                    # Re-search report for retry (critical fix!)
                    self.navigation.search_report(self.report_code)

                    # Update practice object for retry
                    practice.invoice_downloaded = invoice_ok
                    practice.service_downloaded = service_ok

            if invoice_ok and service_ok:
                if self.circuit_breaker:
                    self.circuit_breaker.record_success()
            else:
                log(f"All retries failed for {practice.practice_name}", "error")
                if self.circuit_breaker:
                    self.circuit_breaker.record_failure(practice.practice_name)  # may raise CircuitBreakerTripped

        log("AGE24 Module completed")


# ==========================================================
# FIN18 MODULE - FIXED
# ==========================================================

class FIN18Module:
    """Handles FIN18 report processing - ORIGINAL FILENAME."""

    def __init__(self, driver_manager: ChromeDriverManager,
                 navigation: NavigationModule,
                 report_ops: ReportOperations,
                 download_module: DownloadModule,
                 excel: ExcelManager,
                 auto_fix: Optional['AutoFixEngine'] = None,
                 circuit_breaker: Optional[CircuitBreaker] = None,
                 report_code: str = "FIN18"):
        self.driver_manager = driver_manager
        self.navigation = navigation
        self.report_ops = report_ops
        self.download_module = download_module
        self.excel = excel
        self.driver = driver_manager.driver
        self.auto_fix = auto_fix
        self.circuit_breaker = circuit_breaker
        self.last_exception: Optional[Exception] = None
        self.report_code = report_code  # search term - comes from Excel 'Type' column, not hardcoded

    def prepare(self, practice: str, from_date: str, to_date: str):
        """Prepare FIN18 criteria.

        IMPORTANT: always re-search "FIN18" first - see the AGE24.prepare()
        comment for why. This is called once per practice, so without the
        re-search, practice #2+ would try to select a clinic on the
        RESULTS page left behind by practice #1's download.
        """
        self.navigation.search_report(self.report_code)

        self.report_ops.select_practice(practice)
        self.report_ops.verify_practice(practice)
        self.report_ops.open_criteria()
        self.report_ops.enter_date_range(from_date, to_date)

        log("FIN18 criteria ready")

    def process_practice(self, practice: FIN18Practice) -> bool:
        """Process single FIN18 practice - ORIGINAL FILENAME."""

        log("=" * 80)
        log(f"FIN18 Practice: {practice.practice_name}")
        log(f"Date Range: {practice.from_date} to {practice.to_date}")
        log("=" * 80)

        try:
            self.prepare(practice.practice_name, practice.from_date, practice.to_date)

            # Download with ORIGINAL filename (keep_original_name=True)
            self.download_module.download_report(
                practice.practice_name, 
                Config.FIN18_SUFFIX,
                keep_original_name=True  # Keep original downloaded filename
            )

            self.excel.update_status("FIN18", practice.row_number, 5, Status.DOWNLOADED)
            state_manager.mark_completed(f"FIN18_{practice.practice_name}")

            log(f"FIN18 completed for {practice.practice_name}")
            return True

        except Exception as e:
            log(f"FIN18 failed for {practice.practice_name}: {e}", "error")
            traceback.print_exc()
            self.last_exception = e

            self.driver_manager.save_screenshot(f"FIN18_{practice.practice_name}")
            self.driver_manager.close_extra_windows()
            self.driver_manager.reset_frame()

            self.excel.update_status("FIN18", practice.row_number, 5, Status.FAILED)
            state_manager.mark_failed(f"FIN18_{practice.practice_name}")

            return False

    def run(self):
        """Execute FIN18 module."""
        state_manager.set_module("FIN18")

        log("=" * 80)
        log("Starting FIN18 Module (Original Filename)")
        log("=" * 80)

        # Navigate to FIN18
        self.navigation.search_report(self.report_code)

        # Read practices
        practices = self.excel.read_fin18_practices()

        for practice in practices:
            if not Config.FORCE_REPROCESS:
                if practice.status == "downloaded":
                    log(f"Skipping (already downloaded): {practice.practice_name}")
                    continue

                if state_manager.is_completed(f"FIN18_{practice.practice_name}"):
                    log(f"Skipping (completed in previous run): {practice.practice_name}")
                    self.excel.update_status("FIN18", practice.row_number, 5, Status.DOWNLOADED)
                    continue

            success = False
            self.last_exception = None
            for attempt in range(1, Config.MAX_RETRIES + 1):
                log(f"Attempt {attempt}/{Config.MAX_RETRIES}")

                if self.process_practice(practice):
                    success = True
                    break

                if attempt < Config.MAX_RETRIES:
                    if self.auto_fix and self.last_exception:
                        self.auto_fix.diagnose_and_fix(
                            self.last_exception, context=f"FIN18_{practice.practice_name}"
                        )
                    time.sleep(Config.RETRY_DELAY * attempt)
                    self.navigation.search_report(self.report_code)

            if success:
                if self.circuit_breaker:
                    self.circuit_breaker.record_success()
            else:
                log(f"All retries failed for {practice.practice_name}", "error")
                if self.circuit_breaker:
                    self.circuit_breaker.record_failure(practice.practice_name)  # may raise CircuitBreakerTripped

        log("FIN18 Module completed")


# ==========================================================
# AUTO-FIX ENGINE
# ==========================================================

class AutoFixEngine:
    """Advanced auto-fix flows for common errors."""

    def __init__(self, driver_manager: ChromeDriverManager):
        self.driver_manager = driver_manager
        self.driver = driver_manager.driver
        self.fixes_applied = []

    def diagnose_and_fix(self, error: Exception, context: str = "") -> bool:
        """Diagnose error and apply appropriate fix."""
        error_str = str(error).lower()

        log(f"AutoFix diagnosing: {error_str} in {context}", "warning")

        # Fix 1: Stale element reference
        if "stale" in error_str:
            return self._fix_stale_element()

        # Fix 2: Element not interactable / click intercepted
        if "interactable" in error_str or "intercepted" in error_str:
            return self._fix_element_blocked()

        # Fix 3: Frame issues
        if "frame" in error_str or "no such frame" in error_str:
            return self._fix_frame_issue()

        # Fix 4: Session timeout / disconnected
        if "session" in error_str or "disconnected" in error_str:
            return self._fix_session_issue()

        # Fix 5: Timeout
        if "timeout" in error_str:
            return self._fix_timeout()

        # Fix 6: Practice not found / dropdown issue
        if "practice" in error_str and ("not found" in error_str or "dropdown" in error_str):
            return self._fix_practice_dropdown()

        log("No specific auto-fix available for this error", "warning")
        return False

    def _fix_stale_element(self) -> bool:
        """Fix stale element by waiting and refreshing reference."""
        log("AutoFix: Stale element - waiting and refreshing...")
        time.sleep(2)
        self.driver_manager.reset_frame()
        time.sleep(1)
        self.fixes_applied.append("stale_element_refresh")
        return True

    def _fix_element_blocked(self) -> bool:
        """Fix element blocked by popup or overlay."""
        log("AutoFix: Element blocked - closing popups and scrolling...")
        self.driver_manager.close_extra_windows()
        self.driver_manager.reset_frame()

        # Try to close any overlays
        try:
            self.driver.execute_script("""
                var overlays = document.querySelectorAll('.modal, .overlay, .popup, [class*="modal"], [class*="overlay"]');
                overlays.forEach(function(o) { o.style.display = 'none'; });
            """)
        except:
            pass

        time.sleep(1)
        self.fixes_applied.append("element_blocked_overlay")
        return True

    def _fix_frame_issue(self) -> bool:
        """Fix frame switching issues."""
        log("AutoFix: Frame issue - hard reset...")
        self.driver_manager.hard_reset()
        self.fixes_applied.append("frame_hard_reset")
        return True

    def _fix_session_issue(self) -> bool:
        """Fix session disconnection."""
        log("AutoFix: Session issue - cannot auto-fix, needs re-login", "error")
        self.fixes_applied.append("session_relogin_needed")
        return False  # Cannot auto-fix, needs manual intervention

    def _fix_timeout(self) -> bool:
        """Fix timeout by waiting longer."""
        log("AutoFix: Timeout - increasing wait time...")
        time.sleep(5)
        self.fixes_applied.append("timeout_extended_wait")
        return True

    def _fix_practice_dropdown(self) -> bool:
        """Fix practice dropdown not loading."""
        log("AutoFix: Practice dropdown - re-searching report...")
        self.driver_manager.reset_frame()
        time.sleep(2)
        self.fixes_applied.append("practice_dropdown_research")
        return True

    def get_fix_summary(self) -> str:
        """Get summary of fixes applied."""
        if not self.fixes_applied:
            return "No auto-fixes applied"
        return f"Auto-fixes applied: {', '.join(self.fixes_applied)}"


# ==========================================================
# MAIN CONTROLLER
# ==========================================================

class AutomationController:
    """Main automation orchestrator."""

    def __init__(self):
        self.driver_manager: Optional[ChromeDriverManager] = None
        self.excel: Optional[ExcelManager] = None
        self.login: Optional[LoginModule] = None
        self.navigation: Optional[NavigationModule] = None
        self.report_ops: Optional[ReportOperations] = None
        self.download_module: Optional[DownloadModule] = None
        self.age24: Optional[AGE24Module] = None
        self.fin18: Optional[FIN18Module] = None
        self.auto_fix: Optional[AutoFixEngine] = None
        self.circuit_breaker: Optional[CircuitBreaker] = None

    def initialize(self):
        """Initialize all components."""
        log("=" * 80)
        log("AR Automation Initializing (Advanced Auto-Fix v6)")
        log("=" * 80)

        # Ensure directories
        Config.ensure_directories()

        # Load Excel
        self.excel = ExcelManager(Config.EXCEL_FILE).load()

        # Read credentials before launching browser (fail fast)
        credentials = self.excel.read_login()
        log(f"Credentials loaded for: {credentials.username}")

        # Create driver
        self.driver_manager = ChromeDriverManager()
        driver = self.driver_manager.create_driver()
        self.driver_manager.wait_utils = WaitUtils(driver)

        # Initialize modules
        wait_utils = self.driver_manager.wait_utils

        self.login = LoginModule(driver, wait_utils)
        self.navigation = NavigationModule(driver, wait_utils)
        self.report_ops = ReportOperations(driver, wait_utils)

        download_manager = DownloadManager(Config.DOWNLOAD_DIR)
        self.download_module = DownloadModule(
            self.driver_manager, self.report_ops, download_manager
        )

        self.auto_fix = AutoFixEngine(self.driver_manager)
        self.circuit_breaker = CircuitBreaker()

        # The search-box term ("AGE24"/"FIN18") comes from each sheet's own
        # Type column instead of being hardcoded here.
        age24_report_code = self.excel.get_report_code("AGE24", default="AGE24")
        fin18_report_code = self.excel.get_report_code("FIN18", default="FIN18")

        self.age24 = AGE24Module(
            self.driver_manager, self.navigation, self.report_ops,
            self.download_module, self.excel,
            auto_fix=self.auto_fix, circuit_breaker=self.circuit_breaker,
            report_code=age24_report_code
        )

        self.fin18 = FIN18Module(
            self.driver_manager, self.navigation, self.report_ops,
            self.download_module, self.excel,
            auto_fix=self.auto_fix, circuit_breaker=self.circuit_breaker,
            report_code=fin18_report_code
        )

        # Login
        if not self.login.login(credentials):
            raise Exception("Login failed")

        notification_manager.notify_start()

        log("Initialization complete")

    def run(self):
        """Execute full automation."""
        try:
            self.initialize()

            # Open Administration
            self.navigation.open_administration()

            # Run AGE24
            self.age24.run()

            # Run FIN18
            self.fin18.run()

            # Save workbook
            self.excel.save()

            log("=" * 80)
            log("Automation Completed Successfully")
            if self.auto_fix:
                log(self.auto_fix.get_fix_summary())
            log("=" * 80)

            notification_manager.notify_run_complete(self._build_summary())

        except CircuitBreakerTripped as e:
            log(f"Circuit breaker tripped - stopping run early: {e}", "critical")
            self.excel.save()

            screenshot = None
            if self.driver_manager:
                screenshot = self.driver_manager.save_screenshot("Circuit_Breaker_Tripped")

            notification_manager.notify_fatal_error(e, screenshot=screenshot)
            raise

        except Exception as e:
            log(f"Fatal Error: {e}", "critical")
            traceback.print_exc()

            # Try auto-fix
            if self.auto_fix:
                fixed = self.auto_fix.diagnose_and_fix(e, "main_controller")
                if fixed:
                    log("Auto-fix applied, attempting recovery...", "warning")
                    # Could implement recovery logic here

            screenshot = None
            if self.driver_manager:
                screenshot = self.driver_manager.save_screenshot("Fatal_Error")

            notification_manager.notify_fatal_error(e, screenshot=screenshot)

            raise

        finally:
            self.cleanup()

    def _build_summary(self) -> Dict[str, Any]:
        """Build a run summary dict from state manager for notifications."""
        return {
            "completed": list(state_manager.state.completed_practices),
            "failed": list(state_manager.state.failed_practices),
            "partial": list(state_manager.state.partial_practices),
        }

    def cleanup(self):
        """Cleanup resources."""
        log("Cleaning up...")

        if self.excel:
            try:
                self.excel.save()
                self.excel.close()
            except Exception as e:
                log(f"Excel cleanup error: {e}", "warning")

        if self.driver_manager:
            self.driver_manager.quit()

        state_manager.save()
        log("Cleanup complete")


# ==========================================================
# ENTRY POINT
# ==========================================================

def main():
    """Main entry point."""
    controller = AutomationController()

    try:    
        controller.run()
    except KeyboardInterrupt:
        log("Automation interrupted by user", "warning")
        controller.cleanup()
        sys.exit(1)
    except Exception as e:
        log(f"Unhandled exception: {e}", "critical")
        sys.exit(1)


if __name__ == "__main__":
    main()